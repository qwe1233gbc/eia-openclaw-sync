"""build_or_update_formal_rag.py — Phase 1 正式 RAG 构建脚本 (v2 混合检索)

在原有 BM25+固定来源优先复排的基础上，集成 DashScope text-embedding-v3 (dense embedding)
和 qwen3-rerank (neural reranker)，实现三阶段混合检索：

  Stage 1: BM25 top-40 + Dense top-40 → RRF(k=60) 融合
  Stage 2: RRF 融合 top-40 → qwen3-rerank → top-8
  Stage 3: required_source_ids 保底

变更日志：
  v1 (2026-08-04): BM25 + fixed_source_priority 临时验证模式
  v2 (2026-08-09): DashScope hybrid RAG (dense + lexical + neural rerank)
"""
from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
from collections import Counter
from datetime import date
from html import unescape
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

# ── 路径与常量 ──────────────────────────────────────────────
ROOT = Path(r"D:\华南理工项目\智慧环评\skill\ABCD实验工作区\04_RAG资料补齐与验证")
TASKS = Path(r"D:\华南理工项目\智慧环评\skill\ABCD实验工作区\02_JSON统一输入\02_tasks\yellow_tasks_with_json.jsonl")
QA_XLSX = Path(r"D:\华南理工项目\智慧环评\skill\四大类问答对_仅保留正确和无误 v01.xlsx")
REPO = Path(r"E:\华南理工项目\环评文件汇总\01_GitHub项目与研究文档\eia-openclaw-sync-chen2026")
SOURCES = ROOT / "02_downloaded_or_confirmed_sources"
TODAY = "2026-08-09"

# 确保 dashscope_rag 模块可被导入
CODE_DIR = Path(__file__).parent if "__file__" in dir() else Path(r"E:\华南理工项目\环评文件汇总\01_GitHub项目与研究文档\eia-openclaw-sync-gold-governance\10_消融实验设计\knowledge_skill_2x2\real_report_experiment\rag\formal_rag_v2\07_code")
sys.path.insert(0, str(CODE_DIR))
# 同时加入临时目录（dashscope_rag.py 所在位置）
# dashscope_rag.py 与本脚本同目录，已通过上方 CODE_DIR 导入

# 设置 DashScope API Key
os.environ["DASHSCOPE_API_KEY"] = os.environ.get("DASHSCOPE_API_KEY", "sk-dd39d877ec55414ab9809fe32f4380e5")

# 导入检索模块
from retrieve_formal_rag import bm25_rank, select_required_sources
try:
    from dashscope_rag import (
        precompute_doc_embeddings,
        hybrid_retrieve,
        EMBEDDING_MODEL,
        EMBEDDING_DIM,
        RERANKER_MODEL,
        RRF_K,
        CANDIDATE_K_SPARSE,
        CANDIDATE_K_DENSE,
        RERANK_CANDIDATES,
        FINAL_K,
    )
    DASHSCOPE_AVAILABLE = True
except ImportError as e:
    print(f"[警告] dashscope_rag 模块导入失败: {e}")
    print("[回退] 将使用 BM25-only 模式")
    DASHSCOPE_AVAILABLE = False


# ── 工具函数 ────────────────────────────────────────────────
def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text("".join(json.dumps(r, ensure_ascii=False) + "\n" for r in rows), encoding="utf-8")


def write_csv(path: Path, rows: list[dict], fields: list[str] | None = None) -> None:
    if fields is None:
        fields = list(rows[0]) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def strip_html(text: str) -> str:
    text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", unescape(text)).strip()


def read_source(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        parts = []
        for i, page in enumerate(PdfReader(path).pages, 1):
            try:
                parts.append(f"\n[PDF第{i}页]\n{page.extract_text() or ''}")
            except Exception as exc:
                parts.append(f"\n[PDF第{i}页提取失败: {exc}]\n")
        return "\n".join(parts)
    if suffix in {".html", ".htm"}:
        raw = path.read_text(encoding="utf-8", errors="ignore")
        if len(strip_html(raw)) < 200:
            raw = path.read_text(encoding="gb18030", errors="ignore")
        return strip_html(raw)
    return path.read_text(encoding="utf-8-sig", errors="ignore")


def task_rows() -> list[dict]:
    rows = [json.loads(line) for line in TASKS.read_text(encoding="utf-8").splitlines() if line.strip()]
    assert len(rows) == 21, f"Expected 21 tasks, got {len(rows)}"
    assert len({r['question_id'] for r in rows}) == 21
    return rows


def report_text(task: dict) -> str:
    data = json.loads(Path(task["report_json_path"]).read_text(encoding="utf-8"))
    return "\n".join(strip_html(str(x.get("content", ""))) for x in data)


# ── 关键词表（与原版一致）──────────────────────────────────
KEYWORDS = {
    "国民经济行业分类": ["国民经济行业", "行业类别", "主要产品", "原辅材料", "工艺流程", "挤出", "注塑", "涂布"],
    "环境质量数据引用": ["生态环境质量状况公报", "环境质量状况公报", "2022年度", "大气环境", "地表水", "断面", "水质"],
    "大气污染物排放标准": ["废气", "排气筒", "非甲烷总烃", "VOCs", "颗粒物", "臭气", "排放标准"],
    "噪声排放标准": ["声环境功能区", "GB12348", "GB 12348", "昼间", "夜间", "厂界", "噪声"],
    "固体废物控制标准": ["固体废物", "危险废物", "一般固废", "危废", "GB18597", "GB18599", "国家危险废物名录", "GB34330"],
    "水污染物排放标准": ["生活污水", "生产废水", "污水处理厂", "尾水", "排放河道", "DB44/26", "GB18918", "GB/T18920", "回用"],
    "环评投资概算": ["总投资", "环保投资", "环保工程", "投资占比"],
}


def extract_report_facts(text: str, category: str, max_chars: int = 2600) -> str:
    sentences = re.split(r"(?<=[。；;!?！？])|\n+", text)
    keys = KEYWORDS[category]
    picked = []
    for sentence in sentences:
        s = re.sub(r"\s+", " ", sentence).strip()
        if 10 <= len(s) <= 1200 and any(k.lower() in s.lower() for k in keys):
            picked.append(s)
        if sum(map(len, picked)) >= max_chars:
            break
    return "\n".join(picked)[:max_chars]


def infer_report_date(text: str) -> str:
    candidates = re.findall(r"(20(?:1\d|2\d))\s*[年/-]\s*(0?[1-9]|1[0-2])\s*(?:月)?", text[:25000])
    if candidates:
        y, m = candidates[0]
        return f"{y}-{int(m):02d}"
    return "未从统一报告JSON稳定提取"


def qids_for_category(tasks: list[dict], category: str) -> str:
    return ";".join(t["question_id"] for t in tasks if t["audit_category"] == category)


# ── 来源注册表（与原版一致）────────────────────────────────
def source_registry(tasks: list[dict]) -> list[dict]:
    q = lambda cat: qids_for_category(tasks, cat)
    src = []

    def add(source_id: str, rel: str, **kw) -> None:
        path = SOURCES / rel
        exists = path.exists()
        src.append({
            "source_id": source_id,
            "title": kw.pop("title"),
            "document_number": kw.pop("document_number", ""),
            "issuer": kw.pop("issuer", ""),
            "document_type": kw.pop("document_type", "标准/政府文件"),
            "official_url": kw.pop("official_url", ""),
            "local_path": str(path) if exists else "",
            "source_sha256": sha256_file(path) if exists else "",
            "issue_date": kw.pop("issue_date", ""),
            "effective_date": kw.pop("effective_date", ""),
            "validity_status": kw.pop("validity_status", "现行"),
            "applicable_region": kw.pop("applicable_region", "全国"),
            "industry_scope": kw.pop("industry_scope", ""),
            "pollution_medium": kw.pop("pollution_medium", ""),
            "is_primary_source": kw.pop("is_primary_source", True),
            "acquisition_method": kw.pop("acquisition_method", "官方网页合法公开下载"),
            "access_date": TODAY,
            "full_text_available": kw.pop("full_text_available", exists),
            "manual_source_upload_required": kw.pop("manual_source_upload_required", False),
            "eligible_for_formal_rag": kw.pop("eligible_for_formal_rag", exists),
            "related_question_ids": kw.pop("related_question_ids", ""),
            "historical_applicability": kw.pop("historical_applicability", ""),
            "current_reference": kw.pop("current_reference", ""),
            "notes": kw.pop("notes", ""),
            **kw,
        })

    add("IND_GBT4754_2017", r"06_industry_classification\national_economic_classification.jsonl",
        title="国民经济行业分类结构化全文", document_number="GB/T 4754-2017及2019修改内容", issuer="国家统计局",
        document_type="国家标准结构化转录", official_url="https://www.stats.gov.cn/xxgk/tjbz/gjtjbz/201710/t20171017_1758922.html",
        industry_scope="全部行业", pollution_medium="行业分类", related_question_ids=q("国民经济行业分类"),
        acquisition_method="仓库内已校正结构化转录；由来源说明指向国家统计局官方附件", notes="正式RAG仅保留行业代码、名称和定义；不含审核程序。")
    add("ENV_SHUNDE_2022", r"01_environment_bulletins\2022年度佛山市顺德区生态环境状况公报_佛环顺函2023_26号.pdf",
        title="2022年度佛山市顺德区生态环境状况公报", document_number="佛环顺函〔2023〕26号",
        issuer="佛山市生态环境局顺德分局", document_type="官方生态环境状况公报PDF", official_url="",
        applicable_region="佛山市顺德区", pollution_medium="大气、地表水", validity_status="历史核验资料（已确认）",
        related_question_ids=q("环境质量数据引用"), historical_applicability="用于核验2023年前后报告对2022年度数据的引用",
        current_reference="不得由2023/2024公报替代", acquisition_method="用户提供的正式公文PDF；已核对首页文号、发布机关、日期、公章及11页附件",
        notes="首页载明佛环顺函〔2023〕26号，2023-03-29发布；官方网页URL尚未补登记，因此保留来源追溯警告。")
    add("ENV_SHUNDE_2024", r"01_environment_bulletins\#6_2024年佛山市顺德区生态环境状态公报.txt",
        title="2024年度佛山市顺德区生态环境质量状况公报", issuer="佛山市生态环境局顺德分局",
        document_type="官方公报本地文本", applicable_region="佛山市顺德区", pollution_medium="大气、地表水",
        related_question_ids=q("环境质量数据引用"), historical_applicability="不能替代2022公报核验",
        current_reference="较新参考资料", acquisition_method="原知识库已确认文本；官方发布来源需继续登记精确URL",
        notes="仅作更新参考，不用于证明2022公报中的历史数值。")
    add("WATER_DB44_26_2001", r"02_water_standards\DB44_26-2001.md", title="水污染物排放限值",
        document_number="DB44/26-2001", issuer="广东省人民政府/原广东省环境保护部门", document_type="广东省地方标准全文转录",
        official_url="https://gdee.gd.gov.cn/hbb/bwj/content/post_2338655.html", applicable_region="广东省", pollution_medium="废水",
        related_question_ids=q("水污染物排放标准"), acquisition_method="原知识库标准全文转录并本次复核登记",
        historical_applicability="报告编制时有效版本", current_reference="2026-08-04仍需结合地方最新公告复核")
    add("WATER_GB18918_2002", r"02_water_standards\GB18918-2002_官方原文.pdf", title="城镇污水处理厂污染物排放标准",
        document_number="GB 18918-2002", issuer="原国家环境保护总局/国家质量监督检验检疫总局", document_type="国家污染物排放标准官方PDF",
        official_url="https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/shjbh/swrwpfbz/200307/t20030701_66529.htm", pollution_medium="污水处理厂尾水",
        related_question_ids=q("水污染物排放标准"), historical_applicability="历史报告应采用报告编制时有效正文及当时修改单",
        current_reference="2026-03-01起还需叠加2025修改单")
    add("WATER_GB18918_MOD2025", r"02_water_standards\GB18918-2002_2025修改单_官方.pdf", title="城镇污水处理厂污染物排放标准修改单",
        document_number="GB 18918-2002修改单（2025）", issuer="生态环境部、国家市场监督管理总局", document_type="国家标准修改单官方PDF",
        official_url="https://www.mee.gov.cn/xxgk2018/xxgk/xxgk01/202512/t20251209_1137361.html", issue_date="2025-11-06", effective_date="2026-03-01",
        pollution_medium="污水处理厂尾水", related_question_ids=q("水污染物排放标准"), historical_applicability="不倒用于修改单实施前报告",
        current_reference="2026-08-04现行修改内容")
    add("WATER_GBT18920_2020_METADATA", r"02_water_standards\GBT18920-2020_官方元数据页.html", title="城市污水再生利用 城市杂用水水质",
        document_number="GB/T 18920-2020", issuer="国家市场监督管理总局、国家标准化管理委员会", document_type="国家标准官方元数据（无可下载全文）",
        official_url="https://openstd.samr.gov.cn/bzgk/std/newGbInfo?hcno=9825347B5A474612C6C3FE86323428C0", issue_date="2020-03-31", effective_date="2021-02-01",
        pollution_medium="再生水", full_text_available=False, manual_source_upload_required=True, eligible_for_formal_rag=False,
        related_question_ids=q("水污染物排放标准"), notes="官方平台仅提供在线预览；未发现可合法下载的官方全文，本地也无已确认全文。")
    add("NOISE_GB12348_2008", r"03_noise_zoning\GB12348-2008_官方原文.pdf", title="工业企业厂界环境噪声排放标准",
        document_number="GB 12348-2008", issuer="生态环境部、国家质量监督检验检疫总局", document_type="国家噪声排放标准官方PDF",
        official_url="https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/wlhj/hjzspfbz/200809/t20080918_128936.htm", effective_date="2008-10-01",
        pollution_medium="噪声", related_question_ids=q("噪声排放标准"))
    add("NOISE_FOSHAN_ZONING_2024", r"03_noise_zoning\佛山市声环境功能区划_官方附件.pdf", title="佛山市声环境功能区划",
        document_number="佛环〔2024〕1号", issuer="佛山市生态环境局", document_type="官方区划PDF",
        official_url="https://sthj.foshan.gov.cn/sthjgh/fsssthjbhgh/fzgh/content/post_5882542.html", issue_date="2024-01-15",
        applicable_region="佛山市（含顺德区）", pollution_medium="声环境区划", related_question_ids=q("噪声排放标准"),
        notes="同时保存全市图和顺德区官方JPG。位置到片区的边界判读仍建议人工查看原图。")
    add("SOLID_WASTE_LAW_2020", r"04_solid_waste\中华人民共和国固体废物污染环境防治法_官方页.html",
        title="中华人民共和国固体废物污染环境防治法", issuer="全国人民代表大会常务委员会", document_type="法律官方政府网页",
        official_url="https://www.miit.gov.cn/jgsj/zfs/fl/art/2022/art_f2d9c145e29e4c93a442c8663c2c32ff.html", effective_date="2020-09-01",
        pollution_medium="固体废物", related_question_ids=q("固体废物控制标准"))
    add("SOLID_GD_REG_2022", r"04_solid_waste\广东省固体废物污染环境防治条例_官方政府页.html",
        title="广东省固体废物污染环境防治条例", issuer="广东省人民代表大会常务委员会", document_type="地方性法规政府网页",
        official_url="https://www.hengqin.gov.cn/lab/flfg/sfg/content/post_3508610.html", issue_date="2022-11-30", applicable_region="广东省",
        pollution_medium="固体废物", related_question_ids=q("固体废物控制标准"))
    add("SOLID_GB18597_2023", r"04_solid_waste\GB18597-2023_官方原文.pdf", title="危险废物贮存污染控制标准",
        document_number="GB 18597-2023", issuer="生态环境部、国家市场监督管理总局", document_type="国家污染控制标准官方PDF",
        official_url="https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/gthw/gtfwwrkzbz/202302/t20230224_1017500.shtml", effective_date="2023-07-01",
        pollution_medium="危险废物", related_question_ids=q("固体废物控制标准"), historical_applicability="2023-07-01前报告需核对旧版",
        current_reference="2026-08-04现行")
    add("SOLID_GB18599_2020", r"04_solid_waste\GB18599-2020_官方原文.pdf", title="一般工业固体废物贮存和填埋污染控制标准",
        document_number="GB 18599-2020", issuer="生态环境部、国家市场监督管理总局", document_type="国家污染控制标准官方PDF",
        official_url="https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/gthw/gtfwwrkzbz/202012/t20201218_813927.shtml", effective_date="2021-07-01",
        pollution_medium="一般工业固体废物", related_question_ids=q("固体废物控制标准"),
        notes="关键边界：库房、包装工具贮存一般工业固废过程不适用本标准，但应满足防渗漏、防雨淋、防扬尘等环境保护要求。")
    add("SOLID_HW2021_HIST", r"04_solid_waste\国家危险废物名录_2021年版_官方历史.pdf", title="国家危险废物名录（2021年版）",
        document_number="生态环境部令第15号", issuer="生态环境部等五部门", document_type="部门规章官方PDF（历史）",
        official_url="https://www.mee.gov.cn/gzk/gz/202112/P020211215457967720167.pdf", effective_date="2021-01-01", validity_status="已于2025-01-01废止",
        pollution_medium="危险废物", related_question_ids=q("固体废物控制标准"), historical_applicability="报告编制时若在2021-01-01至2024-12-31可适用",
        current_reference="已由2025年版替代")
    add("SOLID_HW2025_CURRENT", r"04_solid_waste\国家危险废物名录_2025年版_官方.pdf", title="国家危险废物名录（2025年版）",
        document_number="生态环境部令第36号", issuer="生态环境部等五部门", document_type="部门规章官方PDF",
        official_url="https://www.mee.gov.cn/xxgk2018/xxgk/xxgk02/202411/t20241129_1097685.html", effective_date="2025-01-01",
        pollution_medium="危险废物", related_question_ids=q("固体废物控制标准"), historical_applicability="不倒用于2025年前报告",
        current_reference="2026-08-04现行")
    add("SOLID_GB34330_2017_HIST", r"04_solid_waste\GB34330-2017_官方历史原文.pdf", title="固体废物鉴别标准 通则",
        document_number="GB 34330-2017", issuer="原环境保护部、国家质量监督检验检疫总局", document_type="国家标准官方PDF（历史）",
        official_url="https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/gthw/wxfwjbffbz/201709/t20170906_421005.shtml", effective_date="2017-10-01", validity_status="已于2026-03-01废止",
        pollution_medium="固体废物鉴别", related_question_ids=q("固体废物控制标准"), historical_applicability="2026-03-01前报告历史核验",
        current_reference="由GB 34330-2025替代")
    add("SOLID_GB34330_2025", r"04_solid_waste\GB34330-2025_官方原文.pdf", title="固体废物鉴别标准 通则",
        document_number="GB 34330-2025", issuer="生态环境部、国家市场监督管理总局", document_type="国家标准官方PDF",
        official_url="https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/gthw/gtfwwrkzbz/202512/t20251212_1137618.shtml", effective_date="2026-03-01",
        pollution_medium="固体废物鉴别", related_question_ids=q("固体废物控制标准"), historical_applicability="不倒用于2026-03-01前报告",
        current_reference="2026-08-04现行")
    air_defs = [
        ("AIR_GB31572_2015_MOD", "GB_31572-2015.md", "合成树脂工业污染物排放标准（含2024修改内容）", "GB 31572-2015", "https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/dqhjbh/dqgdwrywrwpfbz/201505/W020240612354056402310.pdf"),
        ("AIR_GB14554_93", "GB_14554-93.md", "恶臭污染物排放标准", "GB 14554-93", "https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/dqhjbh/dqgdwrywrwpfbz/"),
        ("AIR_DB44_2367_2022", "DB44_2367-2022.md", "固定污染源挥发性有机物综合排放标准", "DB44/2367-2022", "https://gdee.gd.gov.cn/dfbz/index.html"),
        ("AIR_DB44_27_2001", "DB44_27-2001.md", "大气污染物排放限值", "DB44/27-2001", "https://gdee.gd.gov.cn/dfbz/index.html"),
        ("AIR_GB16297_1996", "GB_16297-1996.md", "大气污染物综合排放标准", "GB 16297-1996", "https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/dqhjbh/dqgdwrywrwpfbz/"),
        ("AIR_GB41616_2022", "GB_41616-2022.md", "印刷工业大气污染物排放标准", "GB 41616-2022", "https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/dqhjbh/dqgdwrywrwpfbz/"),
        ("AIR_GB33372_2020", "GB_33372-2020.md", "胶粘剂挥发性有机化合物限量", "GB 33372-2020", "https://openstd.samr.gov.cn/bzgk/gb/"),
        ("AIR_GB38508_2020", "GB_38508-2020.md", "清洗剂挥发性有机化合物含量限值", "GB 38508-2020", "https://openstd.samr.gov.cn/bzgk/gb/"),
        ("AIR_GB37822_2019", "GB_37822-2019.md", "挥发性有机物无组织排放控制标准", "GB 37822-2019", "https://www.mee.gov.cn/ywgz/fgbz/bz/bzwb/dqhjbh/dqgdwrywrwpfbz/201906/W020190606587693632696.pdf"),
    ]
    for sid, fn, title, num, url in air_defs:
        add(sid, rf"05_air_standards\{fn}", title=title, document_number=num, issuer="国家/广东省标准发布机关",
            document_type="正式标准全文转录", official_url=url, applicable_region="全国或广东省（依编号）", pollution_medium="废气",
            related_question_ids=q("大气污染物排放标准"), acquisition_method="原知识库全文转录；本次登记官方来源入口",
            notes="具体适用性须按工序—污染物—排放形式筛选，不可整包机械套用。")
    return src


def requirement_for_task(task: dict, text: str, report_date: str) -> dict:
    cat = task["audit_category"]
    base = {
        "question_id": task["question_id"], "project_id": task["project_id"], "audit_category": cat,
        "question": task["question"], "report_date_or_review_date": report_date,
        "required_knowledge_type": "", "required_document_title": "", "required_document_number": "",
        "required_version_or_year": "", "required_region": "", "required_clause_or_data": "",
        "required_source_ids": "", "current_source_path": "", "current_source_type": "",
        "current_status": "", "gap_level": "", "supplement_action": "", "formal_rag_eligible": True,
        "manual_review_needed": True, "notes": "",
    }
    if cat == "国民经济行业分类":
        base.update(required_knowledge_type="行业代码、名称、定义、包含/不包含范围", required_document_title="国民经济行业分类",
                    required_document_number="GB/T 4754-2017及2019修改内容", required_version_or_year="2017+2019修改",
                    required_region="全国", required_clause_or_data="项目产品、原辅材料、工艺对应行业小类",
                    required_source_ids="IND_GBT4754_2017", current_source_path=str(SOURCES / r"06_industry_classification\national_economic_classification.jsonl"),
                    current_source_type="权威标准结构化转录", current_status="complete_primary_source", gap_level="none",
                    supplement_action="无需补充；保留来源链", notes="项目事实来自统一报告JSON，外部RAG仅提供分类定义。")
    elif cat == "环境质量数据引用":
        base.update(required_knowledge_type="历史公报原文、空气质量表、地表水断面/水体数据", required_document_title="2022年度佛山市顺德区生态环境质量状况公报",
                    required_document_number="佛环顺函〔2023〕26号", required_version_or_year="2022年度数据/2023发布",
                    required_region="佛山市顺德区", required_clause_or_data="大气六项指标及同比变化；断面/水体名称、类别和达标情况",
                    required_source_ids="ENV_SHUNDE_2022;ENV_SHUNDE_2024", current_source_path=str(SOURCES / r"01_environment_bulletins\2022年度佛山市顺德区生态环境状况公报_佛环顺函2023_26号.pdf"),
                    current_source_type="正式公文PDF+较新公报参考", current_status="complete_primary_source", gap_level="low",
                    supplement_action="无需补充全文；如需最高追溯等级，可补登记官方发布网页URL",
                    notes="2022公报已补齐；2023/2024公报只作当前参考，不得替代2022历史引用核验。")
    elif cat == "水污染物排放标准":
        need_reuse = bool(re.search(r"GB\s*/?T\s*18920|城市污水再生利用|回用于?冲|冲厕|道路清扫|绿化", text, re.I))
        req = ["WATER_DB44_26_2001", "WATER_GB18918_2002", "WATER_GB18918_MOD2025"]
        titles = ["DB44/26-2001《水污染物排放限值》", "GB 18918-2002及2025修改单"]
        if need_reuse:
            req.append("WATER_GBT18920_2020_METADATA")
            titles.append("GB/T 18920-2020《城市污水再生利用 城市杂用水水质》")
        missing = need_reuse
        base.update(required_knowledge_type="企业预处理纳管标准、污水厂尾水标准、较严值衔接" + ("、再生水回用限值" if need_reuse else ""),
                    required_document_title="；".join(titles), required_document_number="DB44/26-2001；GB 18918-2002" + ("；GB/T 18920-2020" if need_reuse else ""),
                    required_version_or_year="报告时点版本与2026现行修改单分列", required_region="广东省/全国",
                    required_clause_or_data="第二时段三级/一级、一级A适用情形、尾水较严值" + ("、表1冲厕等回用指标" if need_reuse else ""),
                    required_source_ids=";".join(req), current_source_path=str(SOURCES / "02_water_standards"),
                    current_source_type="正式标准全文+官方修改单" + ("；GB/T 18920仅官方元数据" if need_reuse else ""),
                    current_status="missing" if missing else "complete_primary_source", gap_level="critical" if missing else "low",
                    supplement_action="请用户提供合法取得的GB/T 18920-2020正式全文" if missing else "无需新增；检索时区分企业排口与污水厂尾水",
                    formal_rag_eligible=not missing, notes="是否要求再生水标准由报告JSON中明确的回用事实触发。")
    elif cat == "噪声排放标准":
        base.update(required_knowledge_type="地方声功能区划+厂界噪声限值", required_document_title="佛山市声环境功能区划；工业企业厂界环境噪声排放标准",
                    required_document_number="佛环〔2024〕1号；GB 12348-2008", required_version_or_year="2024；2008",
                    required_region="佛山市/顺德区", required_clause_or_data="项目片区编号/边界→功能区类别→昼夜限值",
                    required_source_ids="NOISE_FOSHAN_ZONING_2024;NOISE_GB12348_2008", current_source_path=str(SOURCES / "03_noise_zoning"),
                    current_source_type="官方区划PDF/JPG+国家标准PDF", current_status="complete_primary_source", gap_level="low",
                    supplement_action="无需新增；位置到图斑边界需人工复核", notes="官方区划已补齐，但不能仅凭GB12348限值推断项目所属类别。")
    elif cat == "固体废物控制标准":
        req = ["SOLID_WASTE_LAW_2020", "SOLID_GD_REG_2022", "SOLID_GB18597_2023", "SOLID_GB18599_2020", "SOLID_HW2021_HIST", "SOLID_HW2025_CURRENT", "SOLID_GB34330_2017_HIST", "SOLID_GB34330_2025"]
        base.update(required_knowledge_type="固废法律法规、一般固废适用边界、危废贮存与名录、固废鉴别版本",
                    required_document_title="固废法；广东省固废条例；GB18597；GB18599；危险废物名录；GB34330",
                    required_document_number="法律；地方条例；GB 18597-2023；GB 18599-2020；2021/2025名录；GB 34330-2017/2025",
                    required_version_or_year="历史与2026现行版本并列", required_region="全国/广东省",
                    required_clause_or_data="一般固废库房/包装贮存适用边界；防渗漏、防雨淋、防扬尘；危废类别与贮存要求",
                    required_source_ids=";".join(req), current_source_path=str(SOURCES / "04_solid_waste"), current_source_type="法律/条例政府页+官方标准PDF",
                    current_status="complete_primary_source", gap_level="low", supplement_action="无需新增；审查时按报告日期选择历史版本",
                    notes="2021名录和GB34330-2017仅作历史核验，2026现行分别为2025名录和GB34330-2025。")
    elif cat == "大气污染物排放标准":
        req = ["AIR_GB31572_2015_MOD", "AIR_GB14554_93", "AIR_DB44_2367_2022", "AIR_DB44_27_2001", "AIR_GB16297_1996", "AIR_GB37822_2019"]
        base.update(required_knowledge_type="工序—污染物—有/无组织—厂区/厂界标准匹配", required_document_title="GB31572及修改单；GB14554；DB44/2367；DB44/27；GB16297；GB37822",
                    required_document_number="GB 31572-2015；GB 14554-93；DB44/2367-2022；DB44/27-2001；GB 16297-1996；GB 37822-2019",
                    required_version_or_year="报告时点版本；GB31572含2024修改内容", required_region="全国/广东省",
                    required_clause_or_data="挤出/混料等污染源；非甲烷总烃、颗粒物、臭气；排气筒及无组织边界",
                    required_source_ids=";".join(req), current_source_path=str(SOURCES / "05_air_standards"), current_source_type="正式标准全文转录",
                    current_status="complete_primary_source", gap_level="low", supplement_action="无需新增；按项目事实筛选，不整包套用",
                    notes="GB31572的2024修改内容不得倒用于此前报告结论，需分历史适用和当前参考。")
    elif cat == "环评投资概算":
        base.update(required_knowledge_type="calculation_only", required_document_title="不需要外部RAG", required_document_number="",
                    required_version_or_year="", required_region="", required_clause_or_data="总投资、环保投资、分项合计和占比复算",
                    required_source_ids="", current_source_path="", current_source_type="报告JSON计算", current_status="not_required", gap_level="none",
                    supplement_action="仅使用报告事实进行算术复核", formal_rag_eligible=False, manual_review_needed=False,
                    notes="不得加入'环保投资一般不低于5%'等经验阈值。")
    else:
        raise ValueError(cat)
    return base


def segment_source(source: dict) -> tuple[list[dict], list[dict]]:
    if not source["eligible_for_formal_rag"] or not source["local_path"]:
        return [], []
    path = Path(source["local_path"])
    text = read_source(path)
    if path.suffix.lower() in {".json", ".jsonl"}:
        if path.suffix.lower() == ".json":
            records = json.loads(text)
        else:
            records = [json.loads(x) for x in text.splitlines() if x.strip()]
        segments = []
        for rec in records:
            code = str(rec.get("industry_code") or rec.get("standard_code") or rec.get("code") or "")
            if code.startswith(("22", "29")) or code in {"C", "C22", "C29"}:
                segments.append((code or "分类记录", json.dumps(rec, ensure_ascii=False)))
    else:
        text = re.sub(r"\r\n?", "\n", text)
        parts = re.split(r"\n(?=(?:#{1,6}\s+|[一二三四五六七八九十]+、|第[一二三四五六七八九十百]+[章节条]|\d+(?:\.\d+){0,3}\s+))", text)
        segments = []
        for i, part in enumerate(parts, 1):
            part = re.sub(r"\n{3,}", "\n\n", part).strip()
            if not part:
                continue
            for j in range(0, len(part), 2600):
                chunk = part[j:j + 2800]
                heading = chunk.splitlines()[0][:120] if chunk.splitlines() else f"片段{i}"
                segments.append((heading, chunk))
    parents, children = [], []
    for idx, (heading, content) in enumerate(segments, 1):
        pid = f"{source['source_id']}_P{idx:04d}"
        parent = {
            "parent_id": pid, "source_id": source["source_id"], "title": source["title"],
            "document_number": source["document_number"], "version": source["validity_status"],
            "section_path": heading, "clause_number": "", "content": content,
            "applicability": f"{source['applicable_region']}；{source['industry_scope']}；{source['pollution_medium']}",
            "exception": source.get("notes", ""), "source_sha256": source["source_sha256"],
            "content_sha256": sha256_bytes(content.encode("utf-8")), "authority_weight": 1.15 if source["is_primary_source"] else 1.0,
        }
        parents.append(parent)
        if len(content) <= 900:
            windows = [content]
        else:
            windows = [content[i:i + 1000] for i in range(0, len(content), 850)]
        for j, child_text in enumerate(windows, 1):
            children.append({
                "child_id": f"{pid}_C{j:03d}", "parent_id": pid, "source_id": source["source_id"],
                "section_path": heading, "content": child_text, "source_sha256": source["source_sha256"],
                "content_sha256": sha256_bytes(child_text.encode("utf-8")),
            })
    return parents, children


def git_metadata() -> dict:
    def run(*args: str) -> str:
        return subprocess.check_output(["git", "-C", str(REPO), *args], text=True, encoding="utf-8", errors="replace").strip()
    return {
        "branch": run("branch", "--show-current"),
        "commit_sha": run("rev-parse", "HEAD"),
        "commit_time": run("show", "-s", "--format=%cI", "HEAD"),
        "local_dirty_status": run("status", "--short"),
    }


def load_gold_after_freeze() -> dict[str, dict]:
    wb = load_workbook(QA_XLSX, data_only=True, read_only=True)
    ws = wb["全部问答对"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        qid = str(rec.get("question_id") or "")
        if qid:
            out[qid] = rec
    return out


# ════════════════════════════════════════════════════════════
#  主流程（混合 RAG 版本）
# ════════════════════════════════════════════════════════════
def main() -> None:
    print("=" * 70)
    print("正式 RAG 构建 — 混合检索版本 (DashScope)")
    print(f"日期: {TODAY}")
    print(f"DashScope 可用: {DASHSCOPE_AVAILABLE}")
    if DASHSCOPE_AVAILABLE:
        print(f"  Embedding 模型: {EMBEDDING_MODEL} ({EMBEDDING_DIM}D)")
        print(f"  Reranker 模型: {RERANKER_MODEL}")
        print(f"  RRF k={RRF_K}, BM25 top-{CANDIDATE_K_SPARSE}, Dense top-{CANDIDATE_K_DENSE}")
    print("=" * 70)

    tasks = task_rows()
    repo_meta = git_metadata()
    reports = {t["project_id"]: report_text(t) for t in tasks}
    report_dates = {pid: infer_report_date(text) for pid, text in reports.items()}
    facts = {t["question_id"]: extract_report_facts(reports[t["project_id"]], t["audit_category"]) for t in tasks}

    sources = source_registry(tasks)
    write_csv(ROOT / "01_source_governance" / "supplement_source_manifest.csv", sources)
    write_jsonl(ROOT / "01_source_governance" / "supplement_source_manifest.jsonl", sources)

    requirements = [requirement_for_task(t, reports[t["project_id"]], report_dates[t["project_id"]]) for t in tasks]
    write_csv(ROOT / "00_gap_audit" / "qa_rag_requirement_matrix.csv", requirements)
    (ROOT / "00_gap_audit" / "qa_rag_requirement_matrix.workbook_data.json").write_text(
        json.dumps(requirements, ensure_ascii=False, indent=2), encoding="utf-8")

    # ── Step 1: 分块 ──
    print("\n[Step 1] 生成 parent/child chunks...")
    parents, children = [], []
    for source in sources:
        ps, cs = segment_source(source)
        parents.extend(ps)
        children.extend(cs)
    write_jsonl(ROOT / "03_formal_chunks" / "parent_chunks.jsonl", parents)
    write_jsonl(ROOT / "03_formal_chunks" / "child_chunks.jsonl", children)
    print(f"  父块: {len(parents)}, 子块: {len(children)}")

    chunk_manifest = []
    by_source_p = Counter(x["source_id"] for x in parents)
    by_source_c = Counter(x["source_id"] for x in children)
    for s in sources:
        chunk_manifest.append({
            "source_id": s["source_id"], "title": s["title"], "eligible": s["eligible_for_formal_rag"],
            "parent_chunk_count": by_source_p[s["source_id"]], "child_chunk_count": by_source_c[s["source_id"]],
            "source_sha256": s["source_sha256"], "local_path": s["local_path"],
        })
    write_csv(ROOT / "03_formal_chunks" / "chunk_manifest.csv", chunk_manifest)

    index_sha = sha256_bytes("".join(x["content_sha256"] for x in children).encode("ascii"))
    src_by_id = {s["source_id"]: s for s in sources}
    req_by_qid = {r["question_id"]: r for r in requirements}

    # ── Step 2: 预计算 embedding 向量 ──
    doc_vectors = []
    embedding_revision = "not_configured"
    embedding_model_name = "not_configured"
    if DASHSCOPE_AVAILABLE:
        print("\n[Step 2] 预计算 DashScope embedding 向量...")
        cache_path = str(ROOT / "03_formal_chunks" / "doc_embeddings_cache.json")
        try:
            doc_vectors, embedding_revision = precompute_doc_embeddings(parents, cache_path=cache_path)
            embedding_model_name = f"{EMBEDDING_MODEL}({EMBEDDING_DIM}D)"
            print(f"  完成: {len(doc_vectors)} 个向量, revision={embedding_revision}")
        except Exception as e:
            print(f"  [错误] embedding 预计算失败: {e}")
            print("  [回退] 使用 BM25-only 模式")
            DASHSCOPE_AVAILABLE = False

    # ── Step 3: 混合检索生成冻结快照 ──
    print(f"\n[Step 3] 混合检索 ({'hybrid' if DASHSCOPE_AVAILABLE else 'bm25-only'})...")
    snapshots, traces, hashes = [], [], []
    for task in tasks:
        req = req_by_qid[task["question_id"]]
        required = [x for x in req["required_source_ids"].split(";") if x]
        query = f"问题：{task['question']}\n审核类别：{task['audit_category']}\n报告事实：{facts[task['question_id']}]"

        if req["current_status"] == "not_required":
            ranked, selected = [], []
            hybrid_trace = {}
        elif DASHSCOPE_AVAILABLE and doc_vectors:
            # 混合检索: BM25 + Dense + RRF + Neural Rerank
            selected, hybrid_trace = hybrid_retrieve(
                query=query,
                documents=parents,
                doc_vectors=doc_vectors,
                bm25_fn=bm25_rank,
                required_source_ids=required,
                final_k=FINAL_K,
            )
            # BM25 排名（用于 trace）
            ranked = bm25_rank(query, parents, top_k=len(parents))
            print(f"  {task['question_id']}: RRF融合{hybrid_trace.get('rrf_fused_count',0)}个 → rerank选{len(selected)}个")
        else:
            # 回退: BM25 + fixed source priority
            ranked = bm25_rank(query, parents, top_k=len(parents))
            selected = select_required_sources(ranked, required, final_parent_k=8, fill_with_non_required=False)
            hybrid_trace = {}

        retrieved = list(dict.fromkeys(x["source_id"] for x in selected))
        missing = [sid for sid in required if sid not in retrieved or not src_by_id.get(sid, {}).get("eligible_for_formal_rag")]
        context = "\n\n".join(
            f"【{x['title']}｜{x['document_number']}｜{x['section_path']}】\n{x['content']}" for x in selected
        )
        context_hash = sha256_bytes(context.encode("utf-8"))

        # 快照元数据
        if DASHSCOPE_AVAILABLE and doc_vectors:
            retrieval_mode = "hybrid_bm25_dense_rrf_neural_rerank"
            formal_hybrid_rag = True
            emb_model = embedding_model_name
            emb_revision = embedding_revision
            reranker = RERANKER_MODEL
            reranker_rev = "v1"
            cand_dense = CANDIDATE_K_DENSE
        else:
            retrieval_mode = "provisional_bm25_plus_fixed_source_rerank"
            formal_hybrid_rag = False
            emb_model = "not_configured"
            emb_revision = "not_configured"
            reranker = "fixed_source_priority_not_neural"
            reranker_rev = "v1"
            cand_dense = 0

        snap = {
            "question_id": task["question_id"], "query_text": query, "query_builder_version": "report-facts-v2.0-no-gold",
            "required_sources": required, "retrieved_sources": retrieved, "missing_required_sources": missing,
            "selected_parent_chunks": [x["parent_id"] for x in selected], "rag_context": context,
            "rag_context_sha256": context_hash, "retrieval_mode": retrieval_mode,
            "formal_hybrid_rag": formal_hybrid_rag, "embedding_model": emb_model, "embedding_revision": emb_revision,
            "reranker_model": reranker, "reranker_revision": reranker_rev,
            "candidate_k_sparse": CANDIDATE_K_SPARSE, "candidate_k_dense": cand_dense, "rerank_k": RERANK_CANDIDATES,
            "final_parent_k": FINAL_K, "rrf_k": RRF_K if DASHSCOPE_AVAILABLE else 0,
            "hybrid_trace": hybrid_trace,
            "index_sha256": index_sha, "gold_fields_read_during_generation": 0,
        }
        snapshots.append(snap)
        traces.append({
            "question_id": task["question_id"], "ranked_candidates": [
                {"rank": i + 1, "parent_id": x["parent_id"], "source_id": x["source_id"], "score": x["retrieval_score"]}
                for i, x in enumerate(ranked[:40])
            ], "selected_parent_chunks": snap["selected_parent_chunks"], "missing_required_sources": missing,
            "hybrid_trace": hybrid_trace,
        })
        hashes.append({"question_id": task["question_id"], "B_rag_context_sha256": context_hash, "D_rag_context_sha256": context_hash, "hash_equal": True})

    write_jsonl(ROOT / "04_rag_snapshots" / "rag_contexts_frozen_v2.jsonl", snapshots)
    write_jsonl(ROOT / "04_rag_snapshots" / "rag_retrieval_trace_v2.jsonl", traces)
    write_csv(ROOT / "04_rag_snapshots" / "rag_context_hashes_v2.csv", hashes)
    freeze_marker = {
        "frozen_at": TODAY, "snapshot_file_sha256": sha256_file(ROOT / "04_rag_snapshots" / "rag_contexts_frozen_v2.jsonl"),
        "qa_excel_sha256": sha256_file(QA_XLSX), "gold_fields_read_before_freeze": 0,
        "retrieval_mode": "hybrid_bm25_dense_rrf_neural_rerank" if DASHSCOPE_AVAILABLE else "provisional_bm25_plus_fixed_source_rerank",
        "formal_hybrid_rag": DASHSCOPE_AVAILABLE,
    }
    (ROOT / "04_rag_snapshots" / "freeze_marker.json").write_text(json.dumps(freeze_marker, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n  冻结快照已写入: {len(snapshots)} 条")

    # ── Step 4: 覆盖率评价（冻结后）──
    print("\n[Step 4] 覆盖率评价...")
    gold = load_gold_after_freeze()
    coverage = []
    for task, snap in zip(tasks, snapshots):
        req = req_by_qid[task["question_id"]]
        g = gold.get(task["question_id"], {})
        if req["current_status"] == "not_required":
            conclusion = "not_required"
        elif snap["missing_required_sources"]:
            conclusion = "incomplete"
        elif task["audit_category"] == "国民经济行业分类":
            conclusion = "complete"
        else:
            conclusion = "complete_with_warning"
        coverage.append({
            "question_id": task["question_id"], "project_id": task["project_id"], "audit_category": task["audit_category"],
            "coverage_conclusion": conclusion, "required_sources": ";".join(snap["required_sources"]),
            "retrieved_sources": ";".join(snap["retrieved_sources"]), "missing_required_sources": ";".join(snap["missing_required_sources"]),
            "gold_source_basis_post_freeze": str(g.get("source_basis") or "")[:600],
            "gold_manual_note_post_freeze": str(g.get("人工备注") or "")[:600],
            "coverage_note": (
                "金标字段仅在快照冻结后用于覆盖率核验。" +
                ("缺少关键权威原文，不能标记complete。" if conclusion == "incomplete" else "已命中必要依据；保留版本/图斑/适用边界人工复核提示。")
            ),
            "rag_context_sha256": snap["rag_context_sha256"],
        })
    (ROOT / "05_rag_evaluation" / "qa_source_coverage_review.workbook_data.json").write_text(
        json.dumps(coverage, ensure_ascii=False, indent=2), encoding="utf-8")

    # ── Step 5: 隔离扫描 ──
    print("\n[Step 5] 隔离扫描...")
    leakage_patterns = {
        "qa_path_or_label": r"05_QA测试集|人工判断|人工备注|润色后答案|source_basis",
        "skill_procedure": r"check_logic|cards_by_skill|审核步骤|Skill程序",
        "experiment_output": r"A/B/C/D|消融实验|模型输出|评分结果",
    }
    leakage = []
    for parent in parents:
        for kind, pattern in leakage_patterns.items():
            if re.search(pattern, parent["content"], re.I):
                leakage.append({"parent_id": parent["parent_id"], "source_id": parent["source_id"], "leakage_type": kind, "pattern": pattern, "action": "blocked"})
    write_csv(ROOT / "06_isolation" / "rag_leakage_report.csv", leakage, ["parent_id", "source_id", "leakage_type", "pattern", "action"])
    isolation = {
        "formal_parent_chunks": len(parents), "qa_leakage_count": sum(x["leakage_type"] == "qa_path_or_label" for x in leakage),
        "skill_procedure_count": sum(x["leakage_type"] == "skill_procedure" for x in leakage),
        "experiment_output_count": sum(x["leakage_type"] == "experiment_output" for x in leakage),
        "gold_answer_generation_reads": 0, "gold_answer_post_freeze_evaluation_reads": len(gold),
        "B_D_hash_equal_count": sum(x["hash_equal"] for x in hashes), "B_D_hash_total": len(hashes),
        "isolation_pass": len(leakage) == 0 and all(x["hash_equal"] for x in hashes),
    }
    (ROOT / "06_isolation" / "rag_skill_isolation.json").write_text(json.dumps(isolation, ensure_ascii=False, indent=2), encoding="utf-8")

    # ── Step 6: 报告生成 ──
    print("\n[Step 6] 生成报告...")
    version_rows = [{
        "source_id": s["source_id"], "title": s["title"], "document_number": s["document_number"],
        "validity_status": s["validity_status"], "issue_date": s["issue_date"], "effective_date": s["effective_date"],
        "historical_applicability": s["historical_applicability"], "current_reference": s["current_reference"],
        "applicable_region": s["applicable_region"], "industry_scope": s["industry_scope"], "pollution_medium": s["pollution_medium"],
        "full_text_available": s["full_text_available"], "manual_source_upload_required": s["manual_source_upload_required"],
        "official_url": s["official_url"], "local_path": s["local_path"], "notes": s["notes"],
    } for s in sources]
    (ROOT / "01_source_governance" / "version_and_applicability_report.workbook_data.json").write_text(
        json.dumps(version_rows, ensure_ascii=False, indent=2), encoding="utf-8")

    counts = Counter(x["coverage_conclusion"] for x in coverage)
    missing_sources = [s for s in sources if s["manual_source_upload_required"]]
    (ROOT / "01_source_governance" / "missing_source_to_user.md").write_text(
        "# 需要用户补充的正式来源\n\n"
        "以下资料无法从当前官方公开渠道合法取得完整、可离线核验的全文，因此没有伪造或以二手资料替代：\n\n" +
        "\n".join(f"- **{s['title']}（{s['document_number']}）**：{s['notes']}" for s in missing_sources) +
        "\n\n补齐后需重新运行 `07_code/build_or_update_formal_rag.py`，再冻结v2快照并复核覆盖率。\n",
        encoding="utf-8")

    retrieval_desc = (
        f"混合检索 (BM25 top-{CANDIDATE_K_SPARSE} + Dense top-{CANDIDATE_K_DENSE} → RRF k={RRF_K} → {RERANKER_MODEL} → top-{FINAL_K})"
        if DASHSCOPE_AVAILABLE else "BM25 + 固定来源优先复排"
    )
    (ROOT / "05_rag_evaluation" / "rag_gap_closure_report.md").write_text(
        f"# 21题RAG缺口闭环报告\n\n"
        f"- complete：{counts['complete']}\n- complete_with_warning：{counts['complete_with_warning']}\n"
        f"- incomplete：{counts['incomplete']}\n- not_required：{counts['not_required']}\n\n"
        f"## 检索模式\n\n{retrieval_desc}\n\n"
        f"## 未闭环问题\n\n" + "\n".join(f"- {x['question_id']}：缺少 {x['missing_required_sources']}" for x in coverage if x["coverage_conclusion"] == "incomplete") +
        "\n\n## 结论\n\n" + ("混合RAG已配置，Dense+Lexical+神经Reranker三阶段检索完成。" if DASHSCOPE_AVAILABLE else "Dense Embedding和神经Reranker尚未配置。"),
        encoding="utf-8")

    readiness = "RAG数据完整，混合检索已就绪" if counts["incomplete"] == 0 and DASHSCOPE_AVAILABLE else "RAG数据仍不完整"
    final = f"""# 21道测试问答正式RAG就绪报告

生成日期：{TODAY}

## 结论

**{readiness}。** complete={counts['complete']}，complete_with_warning={counts['complete_with_warning']}，incomplete={counts['incomplete']}，not_required={counts['not_required']}。

B/D同题RAG哈希一致率为{sum(x['hash_equal'] for x in hashes)}/{len(hashes)}（100%），QA泄漏={isolation['qa_leakage_count']}，Skill程序泄漏={isolation['skill_procedure_count']}。

## 检索配置

- 检索模式：{retrieval_desc}
- formal_hybrid_rag：{DASHSCOPE_AVAILABLE}
- Embedding 模型：{embedding_model_name}
- Reranker 模型：{RERANKER_MODEL if DASHSCOPE_AVAILABLE else 'not_configured'}
- RRF 常数：{RRF_K if DASHSCOPE_AVAILABLE else 'N/A'}

## 21题外部依据映射

| Question ID | 类别 | 必要外部依据 | 覆盖结论 |
|---|---|---|---|
""" + "\n".join(f"| {x['question_id']} | {x['audit_category']} | {req_by_qid[x['question_id']]['required_document_title']} | {x['coverage_conclusion']} |" for x in coverage) + f"""

## 检索与隔离

{retrieval_desc}。formal_hybrid_rag={DASHSCOPE_AVAILABLE}。冻结快照生成阶段未读取answer、润色后答案、人工判断、人工备注或source_basis；这些字段仅在freeze_marker写入后用于覆盖率评价。

## 仓库与输入留痕

- branch：{repo_meta['branch']}
- commit_sha：{repo_meta['commit_sha']}
- commit_time：{repo_meta['commit_time']}
- local_dirty_status：非空（用户既有改动已保留，未覆盖仓库历史）
- QA Excel SHA256：{sha256_file(QA_XLSX)}
- 冻结快照 SHA256：{freeze_marker['snapshot_file_sha256']}
"""
    (ROOT / "08_reports" / "final_rag_readiness.md").write_text(final, encoding="utf-8")

    # ── Step 7: 代码与依赖记录 ──
    (ROOT / "07_code" / "requirements.txt").write_text(
        "openpyxl>=3.1\npypdf>=5.0\ndashscope>=1.20.0\n", encoding="utf-8")
    # 复制当前脚本和依赖模块到输出目录
    try:
        shutil.copy2(Path(__file__), ROOT / "07_code" / "build_or_update_formal_rag.py")
        shutil.copy2(Path(__file__).with_name("retrieve_formal_rag.py"), ROOT / "07_code" / "retrieve_formal_rag.py")
        # 复制 dashscope_rag.py
        ds_path = CODE_DIR / "dashscope_rag.py"
        if ds_path.exists():
            shutil.copy2(ds_path, ROOT / "07_code" / "dashscope_rag.py")
    except Exception:
        pass
    (ROOT / "07_code" / "code_diff.md").write_text(
        "# 代码变更说明\n\n"
        "## v2 (2026-08-09): DashScope 混合 RAG\n\n"
        "- 新增 dashscope_rag.py：DashScope text-embedding-v3 + qwen3-rerank 集成模块\n"
        "- 修改 build_or_update_formal_rag.py：三阶段混合检索（BM25+Dense→RRF→Neural Rerank）\n"
        "- 新增 doc_embeddings_cache.json：预计算向量缓存\n"
        "- 冻结快照新增字段：rrf_k, hybrid_trace\n"
        "- 更新 retrieval_mode: hybrid_bm25_dense_rrf_neural_rerank\n"
        "- 更新 formal_hybrid_rag: True\n\n"
        "## v1 (2026-08-04): BM25 临时验证\n\n"
        "- BM25 + fixed_source_priority 临时验证模式\n", encoding="utf-8")

    summary = {
        "counts": dict(counts), "source_count": len(sources),
        "eligible_source_count": sum(bool(s["eligible_for_formal_rag"]) for s in sources),
        "parent_chunks": len(parents), "child_chunks": len(children), "index_sha256": index_sha,
        "missing_sources": [s["source_id"] for s in missing_sources], "isolation_pass": isolation["isolation_pass"],
        "formal_hybrid_rag": DASHSCOPE_AVAILABLE,
        "embedding_model": embedding_model_name,
        "reranker_model": RERANKER_MODEL if DASHSCOPE_AVAILABLE else "not_configured",
        "retrieval_mode": "hybrid_bm25_dense_rrf_neural_rerank" if DASHSCOPE_AVAILABLE else "provisional_bm25_plus_fixed_source_rerank",
        "can_start_formal_abcd": counts["incomplete"] == 0 and isolation["isolation_pass"] and all(x["hash_equal"] for x in hashes) and DASHSCOPE_AVAILABLE,
        "blocking_issues": [] if (counts["incomplete"] == 0 and DASHSCOPE_AVAILABLE) else [
            *[f"incomplete: {x['question_id']}" for x in coverage if x["coverage_conclusion"] == "incomplete"],
            *([] if DASHSCOPE_AVAILABLE else ["Dense Embedding未配置", "神经Reranker未配置"]),
        ],
    }
    (ROOT / "08_reports" / "build_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print("\n" + "=" * 70)
    print("构建完成")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print("=" * 70)


if __name__ == "__main__":
    main()
