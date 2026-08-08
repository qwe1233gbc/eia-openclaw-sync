#!/usr/bin/env python3
"""Search the bundled official GB/T 4754—2017 classification notes."""

from __future__ import annotations

import argparse
import re
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore", message="Conditional Formatting extension is not supported.*"
)


SOURCE = (
    Path(__file__).resolve().parent.parent
    / "references"
    / "2017国民经济行业分类注释_按第1号修改单修订.xlsx"
)


def clean(value: object) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def is_code(value: str, length: int) -> bool:
    return bool(re.fullmatch(rf"\d{{{length}}}", value))


def load_records(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    records: list[dict[str, str]] = []
    current_letter = ""
    current_major = ("", "")
    current_middle = ("", "")
    active: dict[str, object] | None = None

    def finish() -> None:
        nonlocal active
        if active is None:
            return
        notes = []
        for item in active.pop("note_lines"):
            if item and (not notes or notes[-1] != item):
                notes.append(item)
        active["notes"] = "\n".join(notes)
        records.append(active)  # type: ignore[arg-type]
        active = None

    for row in sheet.iter_rows(values_only=True):
        values = [clean(value) for value in row[:5]]
        col_a, col_b, _, col_d, _ = values
        if re.fullmatch(r"[A-Z]", col_a):
            finish()
            current_letter = col_a
            current_major = ("", "")
            current_middle = ("", "")
            continue
        if is_code(col_a, 2):
            finish()
            current_major = (col_a, col_d)
            current_middle = ("", "")
            continue
        if is_code(col_a, 3):
            finish()
            current_middle = (col_a, col_d)
            continue
        if is_code(col_b, 4):
            finish()
            active = {
                "letter": current_letter,
                "major_code": current_major[0],
                "major_name": current_major[1],
                "middle_code": current_middle[0],
                "middle_name": current_middle[1],
                "code": col_b,
                "name": col_d,
                "note_lines": [],
            }
            continue
        if active is not None:
            line = " ".join(value for value in values[2:5] if value)
            if line:
                active["note_lines"].append(line)  # type: ignore[index,union-attr]
    finish()
    workbook.close()
    return records


def display(record: dict[str, str]) -> str:
    industry_code = f"{record['letter']}{record['code']}"
    hierarchy = (
        f"{record['major_code']} {record['major_name']} > "
        f"{record['middle_code']} {record['middle_name']}"
    )
    notes = record["notes"] or "（该小类未附展开注释）"
    return f"{industry_code} {record['name']}\n层级：{hierarchy}\n注释：\n{notes}"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="检索国家统计局《2017国民经济行业分类注释》（按第1号修改单修订）"
    )
    parser.add_argument("--code", help="四位小类代码，可写2929或C2929")
    parser.add_argument(
        "--keyword", action="append", default=[], help="产品、工艺、原料或用途关键词，可重复"
    )
    parser.add_argument("--limit", type=int, default=8, help="关键词检索最多返回条数，默认8")
    parser.add_argument("--file", type=Path, default=SOURCE, help="另行指定官方注释xlsx")
    args = parser.parse_args()

    if not args.code and not args.keyword:
        parser.error("至少提供 --code 或 --keyword")
    if not args.file.is_file():
        parser.error(f"找不到分类注释文件：{args.file}")

    records = load_records(args.file)
    results: list[tuple[int, dict[str, str]]] = []

    if args.code:
        code = re.sub(r"^[A-Za-z]", "", args.code.strip())
        results = [(1000, record) for record in records if record["code"] == code]
    else:
        keywords = [keyword.strip().casefold() for keyword in args.keyword if keyword.strip()]
        for record in records:
            name = record["name"].casefold()
            notes = record["notes"].casefold()
            score = 0
            matched = True
            for keyword in keywords:
                if keyword in name:
                    score += 20
                elif keyword in notes:
                    score += 5
                else:
                    matched = False
                    break
            if matched:
                results.append((score, record))
        results.sort(key=lambda item: (-item[0], item[1]["code"]))
        results = results[: max(args.limit, 1)]

    if not results:
        print("未找到匹配小类。请改用最终产品同义词、核心工艺或用途重新检索。")
        return 1

    print(f"依据文件：{args.file.name}")
    for index, (_, record) in enumerate(results, start=1):
        if index > 1:
            print("\n---")
        print(display(record))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
